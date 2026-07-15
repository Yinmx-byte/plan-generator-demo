"""Shared helpers for RDS + OSS maintenance-plan archive storage."""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARCHIVE_HEADERS = [
    "序号", "归档日期", "版本号", "产品类型", "系统名称", "检修动作",
    "方案标题", "网络环境", "实施地点", "所属组织", "所属资源集",
    "检修开始时间", "检修结束时间", "方案提供人", "检修执行人",
    "检修复核人", "安全责任人", "业务影响", "回滚方式",
    "关键参数", "变更摘要", "文件路径",
]

# Actions sorted by length (longest first) for greedy extraction.
MAINTENANCE_ACTIONS = [
    "维护性重启", "读写分离链路切换", "配置变更", "磁盘扩容",
    "回收只读实例", "创建只读实例", "规格变更", "组件扩缩容",
    "回收实例", "创建实例", "白名单变更", "参数调整", "降配",
    "升配", "回收", "创建", "重启", "扩容", "缩容", "变更",
]

PRODUCT_KEYWORDS = [
    ("PolarDB", re.compile(r"polardb", re.I)),
    ("RDS", re.compile(r"\bRDS\b|DRDS", re.I)),
    ("Redis", re.compile(r"redis", re.I)),
    ("SLB", re.compile(r"\bSLB\b|负载均衡", re.I)),
    ("OSS", re.compile(r"\bOSS\b|bucket", re.I)),
    ("K8s", re.compile(r"\bK8s\b|kubernetes|worker", re.I)),
    ("MQ", re.compile(r"\bMQ\b|RocketMQ|GroupID|Topic", re.I)),
    ("ECS", re.compile(r"\bECS\b|云服务器", re.I)),
]

EXCEL_COL_WIDTHS = [5, 12, 7, 10, 18, 10, 35, 10, 22, 14, 14,
                    16, 16, 10, 10, 10, 10, 22, 22, 32, 28, 45]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9)
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
VERSION_FILLS = {
    1: PatternFill("solid", fgColor="FFFFFF"),
    2: PatternFill("solid", fgColor="FFF2CC"),
    3: PatternFill("solid", fgColor="FCE4D6"),
}


def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _maintenance_date(state: dict) -> str:
    """Extract YYYY-MM-DD from schedule_start. Supports ISO and Chinese formats."""
    schedule = str((state or {}).get("schedule_start", "")).strip()
    if not schedule:
        return _today()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", schedule)
    if m:
        return f"{m[1]}-{m[2]}-{m[3]}"
    m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", schedule)
    if m:
        return f"{m[1]}-{int(m[2]):02d}-{int(m[3]):02d}"
    return _today()


def _extract_system_name(instances: str, title: str = "") -> str:
    """Extract the primary system/component name from the instances field."""
    if instances:
        first = re.split(r"\s*[|｜]\s*", instances, maxsplit=1)[0].strip()
        first = re.sub(r"(创建|新增|新建|申请|回收|释放|删除|调整|变更|重启|扩容|缩容|升配|降配).*$", "", first).strip()
        if first:
            return first
    if title:
        clean = title.replace("检修方案", "").strip()
        return clean[:30]
    return "未知系统"


def _extract_action(title: str, instances: str = "", maintenance_type: str = "") -> str:
    """Extract the maintenance action from title, instances, or type."""
    sources = [title, instances]
    for source in sources:
        for action in MAINTENANCE_ACTIONS:
            if action in source:
                return action
    type_map = {
        "配置变更": "配置变更", "组件升级": "组件升级",
        "组件扩缩容": "扩缩容", "数据库变更": "数据库变更",
        "日常维护（原硬件设备）": "日常维护", "其他": "检修",
    }
    return type_map.get(maintenance_type, "检修")


def _extract_product_type(title: str, instances: str = "", tech_params: str = "") -> str:
    """Infer the cloud product type from available text fields."""
    combined = f"{title} {instances} {tech_params}"
    for name, pattern in PRODUCT_KEYWORDS:
        if pattern.search(combined):
            return name
    return "通用"


def _extract_org_and_resource_set(instances: str) -> tuple[str, str]:
    """Parse org and resource set from instances field (pipe-separated)."""
    org, resource_set = "", ""
    if instances:
        parts = [p.strip() for p in re.split(r"\s*[|｜]\s*", instances) if p.strip()]
        if len(parts) >= 2:
            org = parts[1]
        if len(parts) >= 3:
            resource_set = parts[2]
    return org, resource_set


def _extract_impact(document: dict) -> str:
    """Extract a short business impact summary from the risk assessment section."""
    try:
        sections = document.get("document", {}).get("sections", [])
        for sec in sections:
            heading = str(sec.get("heading", "")).replace(" ", "")
            if "风险" not in heading:
                continue
            for block in sec.get("blocks", []):
                if not isinstance(block, dict):
                    continue
                if block.get("type") == "heading" and "影响" in str(block.get("text", "")):
                    continue
                for key in ("text", "items"):
                    val = block.get(key)
                    if isinstance(val, str) and val.strip():
                        return val.strip()[:120]
                    if isinstance(val, list) and val:
                        return str(val[0]).strip()[:120]
        return "待确认"
    except Exception:
        return "待确认"


def _extract_rollback(document: dict) -> str:
    """Extract a short rollback method summary."""
    try:
        sections = document.get("document", {}).get("sections", [])
        for sec in sections:
            heading = str(sec.get("heading", "")).replace(" ", "")
            if "回滚" not in heading:
                continue
            for block in sec.get("blocks", []):
                if not isinstance(block, dict):
                    continue
                if block.get("type") == "heading":
                    continue
                items = block.get("items") or []
                texts = []
                for item in items:
                    if isinstance(item, str):
                        texts.append(item.strip())
                if texts:
                    combined = "; ".join(texts[:2])
                    if "删除新建" in combined:
                        return "删除新建实例"
                    if "恢复原" in combined or "回退原" in combined:
                        return "恢复原配置"
                    if "不涉及回退" in combined or "不涉及回滚" in combined:
                        return "不涉及回退"
                    return combined[:100]
        return "待确认"
    except Exception:
        return "待确认"


def _extract_key_params(state: dict) -> str:
    """Build a concise key parameters summary from tech_params and instances."""
    parts = []
    tech = (state.get("tech_params") or "").strip()
    if tech:
        if len(tech) > 150:
            tech = tech[:147] + "..."
        parts.append(tech)
    instances = (state.get("instances") or "").strip()
    if instances and not parts:
        parts.append(instances[:150])
    if not parts:
        parts.append(state.get("maintenance_type", "检修"))
    return "; ".join(parts)


def _compute_diff_summary(old: dict, new: dict,
                          state_changes: Optional[list[str]] = None) -> str:
    """Generate a one-line change summary — state changes first, content diffs second."""
    parts: list[str] = []

    # Layer 1: precise state-level changes (user intent).
    if state_changes:
        parts.extend(state_changes)
    else:
        # Fallback: compute state changes from old/new plan_data if available
        pass

    # Layer 2: content-level comparison — only supplement when no state changes found.
    if not parts:
        try:
            old_secs = {_norm_heading(s.get("heading", "")): s
                        for s in old.get("document", {}).get("sections", []) if isinstance(s, dict)}
            new_secs = {_norm_heading(s.get("heading", "")): s
                        for s in new.get("document", {}).get("sections", []) if isinstance(s, dict)}

            content_changes = []
            for name in ["实施步骤", "风险评估", "回滚步骤", "现场环境", "实施计划"]:
                o_sec = old_secs.get(name)
                n_sec = new_secs.get(name)
                if not o_sec or not n_sec:
                    continue
                o_blocks = _block_sigs(o_sec)
                n_blocks = _block_sigs(n_sec)
                if o_blocks == n_blocks:
                    continue
                o_set = set(o_blocks)
                n_set = set(n_blocks)
                added = n_set - o_set
                removed = o_set - n_set
                if added and not removed:
                    content_changes.append(f"{name}新增{len(added)}处")
                elif removed and not added:
                    content_changes.append(f"{name}删除{len(removed)}处")
                elif added and removed:
                    content_changes.append(f"{name}修改{len(added) + len(removed)}处")
                else:
                    content_changes.append(f"{name}内容调整")

            if content_changes:
                # No state changes detected — label as LLM variance.
                parts.append("参数未变更（LLM重生成导致措辞微调）")
            else:
                parts.append("未检测到显著变更")
        except Exception:
            return "变更细节无法自动识别，请查看diff报告"
    else:
        # State changes exist — add a brief content supplement.
        try:
            old_secs = {_norm_heading(s.get("heading", "")): s
                        for s in old.get("document", {}).get("sections", []) if isinstance(s, dict)}
            new_secs = {_norm_heading(s.get("heading", "")): s
                        for s in new.get("document", {}).get("sections", []) if isinstance(s, dict)}
            affected = []
            for name in ["实施步骤", "风险评估", "回滚步骤", "现场环境", "实施计划"]:
                o_sec = old_secs.get(name)
                n_sec = new_secs.get(name)
                if not o_sec or not n_sec:
                    continue
                if _block_sigs(o_sec) != _block_sigs(n_sec):
                    affected.append(name)
            if len(affected) >= 4:
                parts.append("多章节内容联动更新")
            elif affected:
                parts.append(f"涉及章节：{'、'.join(affected[:2])}")
        except Exception:
            pass

    return "；".join(parts[:4])


STATE_COMPARE_FIELDS = [
    ("schedule_start", "检修开始时间"),
    ("schedule_end", "检修结束时间"),
    ("schedule_year", "检修年份"),
    ("tech_params", "技术参数"),
    ("instances", "目标实例"),
    ("network", "网络环境"),
    ("location", "实施地点"),
    ("provider", "方案提供人"),
    ("executor", "检修执行人"),
    ("reviewer", "检修复核人"),
    ("security_officer", "安全责任人"),
    ("maintenance_type", "检修类型"),
]


def _compute_state_diff(old_record: dict, new_state: dict) -> list[str]:
    """Compare form state fields between versions. Returns precise change descriptions."""
    changes = []
    new_state = new_state or {}
    for field, label in STATE_COMPARE_FIELDS:
        old_val = str(old_record.get(field, "") or "").strip()
        new_val = str(new_state.get(field, "") or "").strip()
        if old_val == new_val:
            continue
        if not old_val and not new_val:
            continue
        if not old_val:
            changes.append(f"{label}：新增「{new_val[:30]}」")
        elif not new_val:
            changes.append(f"{label}：移除「{old_val[:30]}」")
        else:
            short_old = old_val[:25] + ("…" if len(old_val) > 25 else "")
            short_new = new_val[:25] + ("…" if len(new_val) > 25 else "")
            changes.append(f"{label}：「{short_old}」→「{short_new}」")
    return changes


def _block_sigs(section: dict) -> list[str]:
    """Return a list of stable block signatures for comparison."""
    sigs = []
    for block in section.get("blocks", []):
        if not isinstance(block, dict):
            continue
        btype = block.get("type", "")
        text = str(block.get("text", ""))
        # Collect a canonical representation.
        parts = [btype, text]
        for item in block.get("items", []) or []:
            parts.append(str(item))
        for row in block.get("rows", []) or []:
            if isinstance(row, dict):
                parts.append(json.dumps(row, ensure_ascii=False, sort_keys=True))
            else:
                parts.append(str(row))
        sigs.append("|".join(parts))
    return sigs


def _norm_heading(heading: str) -> str:
    return re.sub(r"\s+", "", heading.lower())


def _generate_diff_markdown(
    old_snapshot: dict, new_snapshot: dict,
    old_version: int, new_version: int, title: str,
    state_changes: Optional[list[str]] = None,
) -> str:
    """Generate a markdown diff report with state changes separated from content diffs."""
    lines = [
        f"# 版本对比: v{old_version} → v{new_version}",
        f"## {title}",
        f"",
        f"**对比时间**: {_now_iso()}",
        f"",
    ]

    # Layer 1: state-level changes (user intent).
    if state_changes:
        lines.append("## 参数变更（用户修改）")
        lines.append("")
        for change in state_changes:
            lines.append(f"- {change}")
        lines.append("")
    else:
        lines.append("## 参数变更（用户修改）")
        lines.append("")
        lines.append("未检测到参数级变更。以下内容差异为 LLM 重生成导致的措辞变化。")
        lines.append("")

    lines.append("## 内容对比（段落级）")
    lines.append("")

    try:
        old_secs = {_norm_heading(s.get("heading", "")): s
                    for s in old_snapshot.get("document", {}).get("sections", []) if isinstance(s, dict)}
        new_secs = {_norm_heading(s.get("heading", "")): s
                    for s in new_snapshot.get("document", {}).get("sections", []) if isinstance(s, dict)}

        all_names = list(dict.fromkeys(list(old_secs.keys()) + list(new_secs.keys())))
        has_changes = False

        for name in all_names:
            o_sec = old_secs.get(name)
            n_sec = new_secs.get(name)
            o_sigs = _block_sigs(o_sec) if o_sec else []
            n_sigs = _block_sigs(n_sec) if n_sec else []

            if o_sigs == n_sigs:
                continue

            has_changes = True
            display_name = o_sec.get("heading", "") if o_sec else (n_sec.get("heading", "") if n_sec else name)

            if not o_sigs:
                lines.append(f"### {display_name} (新增)")
            elif not n_sigs:
                lines.append(f"### {display_name} (已删除)")
            else:
                o_set = set(o_sigs)
                n_set = set(n_sigs)
                removed = o_set - n_set
                added = n_set - o_set
                changed_count = len(removed) + len(added)
                lines.append(f"### {display_name} ({changed_count} 处变更)")
                for sig in list(removed)[:3]:
                    short = sig.split("|", 1)[-1][:120] if "|" in sig else sig[:120]
                    if short.strip():
                        lines.append(f"- ~~{short}~~")
                for sig in list(added)[:3]:
                    short = sig.split("|", 1)[-1][:120] if "|" in sig else sig[:120]
                    if short.strip():
                        lines.append(f"+ **{short}**")

        if not has_changes:
            lines.append("未检测到显著内容变化。")
    except Exception:
        lines.append("差异计算失败，请手动对比两个版本的 plan_snapshot.json。")

    return "\n".join(lines)


def _compute_section_diffs(old_data: dict, new_data: dict) -> list[dict]:
    """Return section-level changes for archive version comparison."""
    diffs: list[dict] = []
    try:
        old_sections = {
            _norm_heading(section["heading"]): section
            for section in old_data.get("document", {}).get("sections", [])
            if isinstance(section, dict) and section.get("heading")
        }
        new_sections = {
            _norm_heading(section["heading"]): section
            for section in new_data.get("document", {}).get("sections", [])
            if isinstance(section, dict) and section.get("heading")
        }
        for name in dict.fromkeys([*old_sections, *new_sections]):
            old_signatures = _block_sigs(old_sections[name]) if name in old_sections else []
            new_signatures = _block_sigs(new_sections[name]) if name in new_sections else []
            if old_signatures == new_signatures:
                status = "unchanged"
            elif not old_signatures:
                status = "added"
            elif not new_signatures:
                status = "removed"
            else:
                status = "modified"
            section = new_sections.get(name) or old_sections.get(name) or {}
            diffs.append({"heading": section.get("heading", ""), "status": status})
    except Exception:
        return []
    return diffs


def write_summary_excel(records: list[dict], output_path: Path) -> None:
    """Build the archive summary workbook from storage-agnostic records."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "检修工作汇总"
    for column, header in enumerate(ARCHIVE_HEADERS, 1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_index, record in enumerate(records, 2):
        values = [
            row_index - 1, record.get("archive_date", ""), f"v{record.get('version', 1)}",
            record.get("product_type", ""), record.get("system_name", ""), record.get("action", ""),
            record.get("title", ""), record.get("network", ""), record.get("location", ""),
            record.get("org", ""), record.get("resource_set", ""), record.get("schedule_start", ""),
            record.get("schedule_end", ""), record.get("provider", ""), record.get("executor", ""),
            record.get("reviewer", ""), record.get("security_officer", ""), record.get("business_impact", ""),
            record.get("rollback_method", ""), record.get("key_params", ""), record.get("change_summary", ""),
            record.get("docx_path", ""),
        ]
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER
            if record.get("version", 1) > 1:
                cell.fill = VERSION_FILLS.get(record.get("version", 1), PatternFill())

    for column, width in enumerate(EXCEL_COL_WIDTHS, 1):
        worksheet.column_dimensions[get_column_letter(column)].width = min(width, 50)
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    workbook.save(str(output_path))


_archive_store = None


def get_archive_store():
    """Return the project's single RDS + OSS archive implementation."""
    global _archive_store
    if _archive_store is None:
        from services.remote_plan_archive import RemoteArchiveStore

        _archive_store = RemoteArchiveStore()
    return _archive_store
